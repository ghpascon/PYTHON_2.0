import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def find_xlsx_file(directory='original'):
    for file in os.listdir(directory):
        if file.endswith('.xlsx'):
            return os.path.join(directory, file)
    return None

def extract_columns_from_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    
    extracted_data = []
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        valid_table_found = False
        
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] is None and row[1] is None:
                if valid_table_found:
                    break
                continue
            
            valid_table_found = True
            extracted_data.append(list(row))
    
    return extracted_data

def save_to_xlsx(data, output_folder):
    # Cria uma pasta de resultado se não existir
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Cria um novo workbook e ativa a planilha
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extraído"
    
    # Define o estilo para células não numéricas (cor cinza claro) e células com números repetidos (cor vermelha)
    fill_gray = PatternFill(start_color="cccccc", end_color="cccccc", fill_type="solid")
    fill_red = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
    
    # Conjunto para rastrear números já vistos
    seen_numbers = set()
    number_positions = {}
    
    # Primeiro, identificamos os números repetidos e suas posições
    for row_index, row in enumerate(data):
        for col_index, value in enumerate(row):
            if isinstance(value, (int, float)):
                if value in seen_numbers:
                    if value not in number_positions:
                        number_positions[value] = []
                    number_positions[value].append((row_index + 1, col_index + 1))
                else:
                    seen_numbers.add(value)
    
    # Adiciona os dados ao workbook e aplica formatação
    for row_index, row in enumerate(data):
        for col_index, value in enumerate(row):
            cell = sheet.cell(row=row_index + 1, column=col_index + 1, value=value)
            # Verifica se o valor não é numérico e pinta a célula de cinza claro
            if not isinstance(value, (int, float)) and value is not None:
                cell.fill = fill_gray
            # Verifica se o número é repetido e pinta a célula de vermelho
            elif isinstance(value, (int, float)) and (row_index + 1, col_index + 1) in [pos for sublist in number_positions.values() for pos in sublist]:
                cell.fill = fill_red

    # Adiciona totais ao final das colunas 1 e 2
    total_col1 = sum(1 for row in data if isinstance(row[0], (int, float)))
    total_col2 = sum(1 for row in data if isinstance(row[1], (int, float)))
    
    last_row = len(data) + 2  # Adiciona duas linhas abaixo da última linha dos dados
    
    # Adiciona o texto "Total" e o total de células numéricas
    celula_metalica= sheet.cell(row=last_row, column=1, value="Total metalica")
    celula_metalica.fill = fill_gray
    sheet.cell(row=last_row + 1, column=1,  value=total_col1)

    celula_nao_metalica = sheet.cell(row=last_row, column=2, value="Total não metalica")
    celula_nao_metalica.fill = fill_gray
    sheet.cell(row=last_row + 1, column=2, value=total_col2)

    # Define o caminho para salvar o arquivo
    output_file = os.path.join(output_folder, 'resultado.xlsx')
    workbook.save(output_file)
    print(f'Arquivo salvo como {output_file}')